import openpyxl

#File-->Workbook-->Sheets-->Rows-->Cells
file = "C:\LevelUp Study\LevelUP School\Random\Data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row #count num of rows in a excel sheet
columns = sheet.max_column #count num of columns in a excel sheet

for r in range(1,rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(r,c).value,end ='  ')
    print()

