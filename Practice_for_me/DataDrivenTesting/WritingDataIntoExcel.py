import openpyxl

# This is for the same data
# file = "C:\LevelUp Study\LevelUP School\Random\Data.xlsx"
# workbook = openpyxl.load_workbook(file)
# # sheet = workbook["Data"] #for multiple sheets -->#get active sheet from excel
# sheet = workbook.active  # get active sheet from excel
#
# for r in range(1, 6):  # 5 rows will create
#     for c in range(1, 4):  # 3 columns will create
#         sheet.cell(r, c).value = "welcome"
#
# workbook.save(file)

# For Multiple data:
file = "C:\LevelUp Study\LevelUP School\Random\Data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "Smith"
sheet.cell(1, 3).value = "Engineer"

sheet.cell(2, 1).value = 567
sheet.cell(2, 2).value = "John"
sheet.cell(2, 3).value = "Manager"

sheet.cell(3, 1).value = 567
sheet.cell(3, 2).value = "David"
sheet.cell(3, 3).value = "Developer"

workbook.save(file)
